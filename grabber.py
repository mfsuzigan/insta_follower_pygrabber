import argparse
from datetime import datetime
from io import BytesIO
import json
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import requests
from urllib.request import urlopen
from urllib.error import HTTPError


GENERAL_WIDTH = 18
PROFILE_IMAGE_WIDTH = 16
PROFILE_IMAGE_HEIGHT = 115
INITIAL_FOLLOWERS_REQUEST_MAX_ID = -1
WORKBOOK_RECORDS_SIZE = 1000

def getImageFromUrl(url):
    try :
        image = Image(BytesIO(urlopen(url).read()))
        image.length = 150
        image.height = 150
        return image

    except HTTPError:
        print(f"Failed getting image from {url}.")

def setupSpreadSheet(sheet):
    sheet["A1"] = "pk"
    sheet["B1"] = "username"
    sheet["C1"] = "full_name"
    sheet["D1"] = "is_private"
    sheet["E1"] = "profile_pic"
    sheet["F1"] = "profile_pic_url"
    
    sheet.column_dimensions['A'].width = GENERAL_WIDTH
    sheet.column_dimensions['B'].width = GENERAL_WIDTH
    sheet.column_dimensions['C'].width = GENERAL_WIDTH
    sheet.column_dimensions['E'].width = PROFILE_IMAGE_WIDTH

def getWorkbookName(pk):
    identifier = datetime.now().strftime("%y%m%d%H%M%S%f")[: -3]
    return f"insta_followers_account_{pk}_output--{identifier}.xlsx"

def getFollowers(pk, xIgAppId, cookie, nextMaxId):
    headers = {"x-ig-app-id": xIgAppId, "cookie": cookie}
    params = None

    if nextMaxId and not nextMaxId == INITIAL_FOLLOWERS_REQUEST_MAX_ID:
        params = {"max_id": nextMaxId}

    response = requests.get(f"https://www.instagram.com/api/v1/friendships/{pk}/followers/", headers=headers, params=params)
    # response = requests.get("http://localhost:8000", headers=headers, params=params)
        
    try:
        # print(response.text)
        return response.json()
    except ValueError:
        print("Failed getting followers.")

def getArgs():
    parser = argparse.ArgumentParser()
    parser.add_argument('username', help='Target Instagram username')
    parser.add_argument('xIgAppId', help='Your Instagram app ID')
    parser.add_argument('cookie', help='Instagram cookie from browser')
    return parser.parse_args()

def writeFollowers(followers, sheet, currentRow):
    for index, element in enumerate(followers["users"]):
        rowNumber = index + currentRow + 1

        sheet[f"A{rowNumber}"] = element["pk"]
        sheet[f"B{rowNumber}"] = element["username"]
        sheet[f"C{rowNumber}"] = element["full_name"]
        sheet[f"D{rowNumber}"] = str(bool(element["is_private"]))
        sheet[f"F{rowNumber}"] = element["profile_pic_url"]

        sheet.row_dimensions[rowNumber].height = PROFILE_IMAGE_HEIGHT
        profileImage = getImageFromUrl(element["profile_pic_url"])

        if profileImage:
            sheet.add_image(profileImage, f"E{rowNumber}")

        print(f"Follower {element['username']} added to spreadsheet")

def getWorkbook():
    workbook = Workbook()
    setupSpreadSheet(workbook.active)
    return workbook

def getIdForUsername(username, xIgAppId):
    response = requests.get(f"https://i.instagram.com/api/v1/users/web_profile_info/?username={username}", headers={"x-ig-app-id": xIgAppId})

    try:
        return response.json()["data"]["user"]["id"]

    except Exception:
        print(f"Failed getting id for username {username}.")

def saveFile(workbook, username):
    filename = getWorkbookName(username)
    workbook.save(filename=filename)
    print(f"\nFile {filename} saved")
    

def main():
    args = getArgs()    

    rowCounter = 0
    followersCounter = 0
    fileSavesCounter = 0

    workbook = getWorkbook()
    nextMaxId = INITIAL_FOLLOWERS_REQUEST_MAX_ID

    pk = getIdForUsername(args.username, args.xIgAppId)

    while True:
        followers = getFollowers(pk, args.xIgAppId, args.cookie, nextMaxId)
        writeFollowers(followers, workbook.active, (rowCounter + 1))

        rowCounter += len(followers["users"])
        followersCounter += len(followers["users"])

        print(f"\nFollowers added: {followersCounter}")

        if followersCounter >= (fileSavesCounter + 1) * WORKBOOK_RECORDS_SIZE:
            saveFile(workbook, args.username)
            fileSavesCounter += 1

            input("\nPress Enter to continue...")

            workbook = getWorkbook()
            rowCounter = 0

        if "next_max_id" in followers:
            nextMaxId = followers["next_max_id"]

        else:
            break        

    saveFile(workbook, args.username)
    print(f"\nFinished.")

if __name__ == '__main__':
    main()